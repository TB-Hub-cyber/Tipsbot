# excel_utils.py
from __future__ import annotations
import io
import re
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook

# ---- Global state (fylls via API:t) ----
KUPONG: List[Dict[str, Any]] = []
FOOTY: Dict[int, Dict[str, Any]] = {}

DATA_SHEET_NAME = "Data"
START_ROW = 2  # Match 1 ligger på rad 2

def reset_state():
    KUPONG.clear()
    FOOTY.clear()

def update_kupong(rows: List[Dict[str, Any]]):
    global KUPONG
    KUPONG = list(rows or [])

def update_footy(matchnr: int, data: Dict[str, Any]):
    FOOTY[int(matchnr)] = data

def _open_wb(path: str):
    return load_workbook(path)

def _sheet(wb, name: str):
    return wb[name] if name in wb.sheetnames else wb.active

def _find_col(ws, exact_header: str) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip() == exact_header:
            return c
    return None

# --------- Exakta rubriker enligt din fil ---------
# Kupong
COL_MATCH         = "Match"
COL_HOME          = "Hemmalag"
COL_AWAY          = "Bortalag"
COL_ODDS_1        = "Odds 1"
COL_ODDS_X        = "Odds X"
COL_ODDS_2        = "Odds 2"
COL_FOLK_1        = "Folkets val 1"
COL_FOLK_X        = "Folkets val X"
COL_FOLK_2        = "Folkets val 2"
COL_SV_1          = "Spelvärde 1"
COL_SV_X          = "Spelvärde X"
COL_SV_2          = "Spelvärde 2"

# Footy
COL_FORM_H        = "Form H (senaste 5)"
COL_FORM_B        = "Form B (senaste 5)"
COL_H2H_5         = "H2H senaste 5"
COL_XG_H_OV       = "xG H (overall)"
COL_XGA_H_OV      = "xGA H (overall)"
COL_XG_B_OV       = "xG B (overall)"
COL_XGA_B_OV      = "xGA B (overall)"
COL_PPG_H_OV      = "PPG H (overall)"
COL_PPG_B_OV      = "PPG B (overall)"

def _set_if_not_none(ws, row: int, col: Optional[int], value):
    if col and value is not None:
        ws.cell(row=row, column=col, value=value)

def write_kupong_into_data_sheet(wb):
    if not KUPONG:
        return
    ws = _sheet(wb, DATA_SHEET_NAME)

    c_match = _find_col(ws, COL_MATCH)
    c_home  = _find_col(ws, COL_HOME)
    c_away  = _find_col(ws, COL_AWAY)
    c_o1    = _find_col(ws, COL_ODDS_1)
    c_ox    = _find_col(ws, COL_ODDS_X)
    c_o2    = _find_col(ws, COL_ODDS_2)
    c_f1    = _find_col(ws, COL_FOLK_1)
    c_fx    = _find_col(ws, COL_FOLK_X)
    c_f2    = _find_col(ws, COL_FOLK_2)
    c_sv1   = _find_col(ws, COL_SV_1)
    c_svx   = _find_col(ws, COL_SV_X)
    c_sv2   = _find_col(ws, COL_SV_2)

    for r in KUPONG:
        try:
            mn = int(r.get("matchnr"))
        except Exception:
            continue
        row = START_ROW - 1 + mn

        _set_if_not_none(ws, row, c_match, mn)
        _set_if_not_none(ws, row, c_home,  r.get("hemmalag"))
        _set_if_not_none(ws, row, c_away,  r.get("bortalag"))
        _set_if_not_none(ws, row, c_o1,    r.get("odds_1"))
        _set_if_not_none(ws, row, c_ox,    r.get("odds_x"))
        _set_if_not_none(ws, row, c_o2,    r.get("odds_2"))
        _set_if_not_none(ws, row, c_f1,    r.get("folk_1"))
        _set_if_not_none(ws, row, c_fx,    r.get("folk_x"))
        _set_if_not_none(ws, row, c_f2,    r.get("folk_2"))
        _set_if_not_none(ws, row, c_sv1,   r.get("spelv_1"))
        _set_if_not_none(ws, row, c_svx,   r.get("spelv_x"))
        _set_if_not_none(ws, row, c_sv2,   r.get("spelv_2"))

def write_footy_into_data_sheet(wb):
    if not FOOTY:
        return
    ws = _sheet(wb, DATA_SHEET_NAME)

    c_form_h = _find_col(ws, COL_FORM_H)
    c_form_b = _find_col(ws, COL_FORM_B)
    c_h2h    = _find_col(ws, COL_H2H_5)
    c_xg_h   = _find_col(ws, COL_XG_H_OV)
    c_xga_h  = _find_col(ws, COL_XGA_H_OV)
    c_xg_b   = _find_col(ws, COL_XG_B_OV)
    c_xga_b  = _find_col(ws, COL_XGA_B_OV)
    c_ppg_h  = _find_col(ws, COL_PPG_H_OV)
    c_ppg_b  = _find_col(ws, COL_PPG_B_OV)

    for mn, d in FOOTY.items():
        row = START_ROW - 1 + int(mn)
        home = d.get("home") or {}
        away = d.get("away") or {}
        h2h  = d.get("h2h_last5") or {}

        _set_if_not_none(ws, row, c_form_h, home.get("form_last5"))
        _set_if_not_none(ws, row, c_form_b, away.get("form_last5"))

        if any(h2h.get(k) is not None for k in ("W","D","L")):
            s = f"W:{h2h.get('W')} D:{h2h.get('D')} L:{h2h.get('L')}"
            _set_if_not_none(ws, row, c_h2h, s)

        _set_if_not_none(ws, row, c_xg_h,  home.get("xg_for"))
        _set_if_not_none(ws, row, c_xga_h, home.get("xg_against"))
        _set_if_not_none(ws, row, c_xg_b,  away.get("xg_for"))
        _set_if_not_none(ws, row, c_xga_b, away.get("xg_against"))
        _set_if_not_none(ws, row, c_ppg_h, home.get("ppg"))
        _set_if_not_none(ws, row, c_ppg_b, away.get("ppg"))

def write_excel_bytes(template_path: str) -> bytes:
    wb = _open_wb(template_path)
    write_kupong_into_data_sheet(wb)
    write_footy_into_data_sheet(wb)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
