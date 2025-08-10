# main.py — FastAPI backend för Tipsbot (Render)
# - Hämtar Stryktips-data från Stryketanalysen (scrape_stryket.fetch_stryket)
# - Exponerar /svenskaspel, /excel, /health, /reset, /debug/state
# - Tjänar /debug/stryket.html via static mount

from fastapi import FastAPI, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Dict, Any
from datetime import datetime
import io
import os
import pathlib

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---- importera vår scraper ----
from scrape_stryket import fetch_stryket, DEBUG_HTML_PATH, STATIC_DIR

# ---------------------------------
# App & CORS
# ---------------------------------
app = FastAPI(title="Tipsbot API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # öppet för enkelhet (kan låsas till pythonistas origin)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# se till att static-katalogen finns och mounta
STATIC_DIR.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Bekväm alias för debug-HTML (så du kan öppna /debug/stryket.html)
@app.get("/debug/stryket.html")
def debug_stryket_html_redirect():
    # Render har redan mountat /static – men vi ger en enkel redirect-lösning
    if DEBUG_HTML_PATH.exists():
        return Response(
            content=(DEBUG_HTML_PATH.read_text(encoding="utf-8")),
            media_type="text/html; charset=utf-8",
            status_code=200
        )
    return Response(content="Ingen debug-HTML sparad ännu.", media_type="text/plain; charset=utf-8", status_code=404)

# ---------------------------------
# State i minne
# ---------------------------------
STATE: Dict[str, Any] = {
    "last_url": None,
    "last_fetch_ts": None,
    "svenskaspel": [],   # list[dict] med 13 matcher
}

# ---------------------------------
# Models
# ---------------------------------
class SvsReq(BaseModel):
    url: str
    debug: bool = False
    footy: List[str] = []   # reserverad; används ej här ännu

# ---------------------------------
# Hjälp: skapa Excel i minne
# ---------------------------------
EXCEL_COLUMNS = [
    ("Matchnr", 8),
    ("Hemmalag", 22),
    ("Bortalag", 22),
    ("Odds_1", 10),
    ("Odds_X", 10),
    ("Odds_2", 10),
    ("Folk_1 (%)", 12),
    ("Folk_X (%)", 12),
    ("Folk_2 (%)", 12),
    ("Spelvärde_1", 12),
    ("Spelvärde_X", 12),
    ("Spelvärde_2", 12),
]

def build_excel(data: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Kupong"

    # rubriker
    for col_idx, (title, width) in enumerate(EXCEL_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=title)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # rader
    row_idx = 2
    for row in data:
        ws.cell(row=row_idx, column=1,  value=row.get("matchnr"))
        ws.cell(row=row_idx, column=2,  value=row.get("hemmalag"))
        ws.cell(row=row_idx, column=3,  value=row.get("bortalag"))
        ws.cell(row=row_idx, column=4,  value=row.get("odds_1"))
        ws.cell(row=row_idx, column=5,  value=row.get("odds_x"))
        ws.cell(row=row_idx, column=6,  value=row.get("odds_2"))
        ws.cell(row=row_idx, column=7,  value=row.get("folk_1"))
        ws.cell(row=row_idx, column=8,  value=row.get("folk_x"))
        ws.cell(row=row_idx, column=9,  value=row.get("folk_2"))
        ws.cell(row=row_idx, column=10, value=row.get("spelv_1"))
        ws.cell(row=row_idx, column=11, value=row.get("spelv_x"))
        ws.cell(row=row_idx, column=12, value=row.get("spelv_2"))
        row_idx += 1

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------------------------------
# Endpoints
# ---------------------------------
@app.get("/health")
def health():
    return {"ok": True, "ts": datetime.utcnow().isoformat()}

@app.post("/reset")
def reset():
    STATE["last_url"] = None
    STATE["last_fetch_ts"] = None
    STATE["svenskaspel"] = []
    # töm debug-html
    try:
        if DEBUG_HTML_PATH.exists():
            DEBUG_HTML_PATH.unlink()
    except Exception:
        pass
    return {"ok": True}

@app.get("/debug/state")
def debug_state():
    return {
        "last_url": STATE["last_url"],
        "last_fetch_ts": STATE["last_fetch_ts"],
        "svenskaspel_rows": len(STATE["svenskaspel"]),
        "debug_html_exists": DEBUG_HTML_PATH.exists(),
    }

@app.post("/svenskaspel")
def svenskaspel(req: SvsReq):
    try:
        # just nu: endast Stryketanalysen-vägen
        if "stryketanalysen.se" in (req.url or ""):
            result = fetch_stryket(req.url, debug=req.debug)
        else:
            raise HTTPException(status_code=400, detail="Okänd URL-källa. Ange Stryketanalysen-URL.")

        rows = result.get("svenskaspel") or []
        if not rows:
            raise HTTPException(status_code=502, detail="Scrape-fel: tomt resultat.")

        # spara i state
        STATE["svenskaspel"] = rows[:13]  # säkerställ 13 rader
        STATE["last_url"] = req.url
        STATE["last_fetch_ts"] = datetime.utcnow().isoformat()

        return {"svenskaspel": STATE["svenskaspel"]}
    except HTTPException:
        raise
    except Exception as e:
        # om scrape_stryket kastar fel hamnar vi här
        raise HTTPException(status_code=502, detail=str(e))

@app.get("/excel")
def excel():
    rows = STATE["svenskaspel"]
    if not rows:
        raise HTTPException(status_code=404, detail="Ingen kupongdata i minnet ännu. Kör /svenskaspel först.")

    content = build_excel(rows)
    filename = f"Stryktipsanalys_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# valfri root
@app.get("/")
def root():
    return {"service": "tipsbot", "status": "ok"}
